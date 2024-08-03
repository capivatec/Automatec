from flask import Flask, render_template, request, redirect, url_for, flash
import openpyxl
import os

app = Flask(__name__)
app.secret_key = '1ae66ef41a4a37e20fae37f4f6dc357b'

# Caminho para o arquivo Excel
FILE_PATH = 'estoque.xlsx'

# Função para carregar a planilha
def load_workbook():
    if not os.path.exists(FILE_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Estoque'
        ws.append(['ID', 'Nome', 'Quantidade'])
        wb.save(FILE_PATH)
    return openpyxl.load_workbook(FILE_PATH)

# Rota para a página inicial
@app.route('/')
def index():
    wb = load_workbook()
    ws = wb.active
    estoque = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        estoque.append({
            'id': row[0],
            'nome': row[1],
            'quantidade': row[2]
        })
    return render_template('index.html', estoque=estoque)

# Rota para adicionar um item ao estoque
@app.route('/add', methods=['POST'])
def add_item():
    nome = request.form['nome']
    quantidade = int(request.form['quantidade'])
    
    wb = load_workbook()
    ws = wb.active
    new_id = ws.max_row  # Assuming IDs are sequential and start from 1
    ws.append([new_id, nome, quantidade])
    wb.save(FILE_PATH)
    
    flash('Item adicionado com sucesso!', 'success')
    return redirect(url_for('index'))

# Rota para editar um item no estoque
@app.route('/edit/<int:item_id>', methods=['POST'])
def edit_item(item_id):
    nome = request.form['nome']
    quantidade = int(request.form['quantidade'])
    
    wb = load_workbook()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == item_id:
            row[1].value = nome
            row[2].value = quantidade
            break
    wb.save(FILE_PATH)
    
    flash('Item atualizado com sucesso!', 'success')
    return redirect(url_for('index'))

# Rota para deletar um item do estoque
@app.route('/delete/<int:item_id>', methods=['POST'])
def delete_item(item_id):
    wb = load_workbook()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == item_id:
            ws.delete_rows(row[0].row)
            break
    wb.save(FILE_PATH)
    
    flash('Item deletado com sucesso!', 'success')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
